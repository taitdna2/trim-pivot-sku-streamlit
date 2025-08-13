# app.py
import os
import csv
from io import BytesIO
from typing import List, Optional
from collections import Counter, defaultdict

import pandas as pd
import streamlit as st

# =========================
# CẤU HÌNH CHUNG
# =========================
st.set_page_config(page_title="CẮT & PIVOT SKU", layout="wide")
st.title("✂️ CẮT FILE & 📊 PIVOT SẢN LƯỢNG/DOANH SỐ THEO KH")
st.caption(
    "Bước 1: cắt cột bằng pandas (giữ nguyên số dòng).  "
    "Bước 2: pivot theo khách hàng — chọn 1 trong 3 chế độ: "
    "Nhanh (pandas), Siêu nhẹ (streaming XLSX), Siêu nhẹ (CSV)."
)

# Cột theo vị trí (0-based): D(3), L(11), M(12), Q(16), R(17), S(18), W(22), Z(25)
COL_INDICES = [3, 11, 12, 16, 17, 18, 22, 25]

# Cột bắt buộc cho PIVOT
REQUIRED = [
    "Tên NPP", "Mã KH", "Tên KH", "Nhóm hàng",
    "Mã SP", "Tên SP", "Tổng Sản lượng (Lẻ)", "Doanh số bán"
]

# Nếu header file “đã cắt” chưa đúng tên → map nhanh theo vị trí
INDEX_TO_REQUIRED = {
    0: "Tên NPP",               # D
    1: "Mã KH",                 # L
    2: "Tên KH",                # M
    3: "Nhóm hàng",             # Q
    4: "Mã SP",                 # R
    5: "Tên SP",                # S
    6: "Tổng Sản lượng (Lẻ)",   # W
    7: "Doanh số bán",          # Z
}

# Alias để auto-map khi pivot streaming
NEEDED_NAMES = {
    "Tên NPP": {"ten npp", "tennpp", "tên npp"},
    "Mã KH": {"ma kh", "mã kh", "ma_kh", "makh", "customer id", "customerid"},
    "Tên KH": {"ten kh", "tên kh", "ten_kh", "tenkh"},
    "Nhóm hàng": {"nhom hang", "nhóm hàng", "nhom_hang"},
    "Mã SP": {"ma sp", "mã sp", "ma_sp", "masp"},
    "Tên SP": {"ten sp", "tên sp", "ten_sp", "tensp"},
    "Tổng Sản lượng (Lẻ)": {
        "tong san luong (le)", "tổng sản lượng (lẻ)",
        "san luong", "sanluong", "tong sl", "tổng sl"
    },
    "Doanh số bán": {"doanh so ban", "doanh số bán", "doanh so", "sales", "revenue"},
}

# =========================
# HELPERS CHUNG
# =========================
def _mode_text(series):
    vals = [str(x).strip() for x in series if str(x).strip() and str(x).strip().lower() != "nan"]
    if not vals:
        return ""
    cnt = Counter(vals)
    mx = max(cnt.values())
    return sorted([v for v, c in cnt.items() if c == mx])[0]

def _normalize_header(h):
    return str(h or "").strip().lower()

# =========================
# BƯỚC 1 — CẮT FILE BẰNG PANDAS (GIỮ NGUYÊN SỐ DÒNG)
# =========================
def cut_with_pandas_keep_rows(file_bytes: bytes, sheet_name: Optional[str]) -> bytes:
    """
    Đọc full sheet (để không mất dòng), chỉ lấy các cột theo COL_INDICES, trả về .xlsx bytes.
    """
    xls = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    real_sheet = sheet_name if (sheet_name and sheet_name in xls.sheet_names) else xls.sheet_names[0]

    df_all = pd.read_excel(xls, sheet_name=real_sheet, dtype=object, keep_default_na=False, engine="openpyxl")

    max_idx = max(COL_INDICES)
    if df_all.shape[1] <= max_idx:
        raise ValueError(f"Sheet '{real_sheet}' chỉ có {df_all.shape[1]} cột, cần tới index {max_idx}. Kiểm tra lại.")

    df_cut = df_all.iloc[:, COL_INDICES]

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df_cut.to_excel(w, index=False, sheet_name="Trimmed")
    return out.getvalue()

# =========================
# BƯỚC 2 — PIVOT (PANDAS)
# =========================
def normalize_after_cut(df: pd.DataFrame, header_row_user: int) -> pd.DataFrame:
    """
    Chuẩn hoá tên cột sau khi cắt (8 cột).
    - header_row_user: dòng tiêu đề (1-based) mà người dùng chọn.
    - Nếu tên chưa ổn → map theo INDEX_TO_REQUIRED.
    """
    if header_row_user != 1:
        buf = BytesIO()
        df.to_excel(buf, index=False, header=False)
        buf.seek(0)
        df = pd.read_excel(buf, header=header_row_user - 1, engine="openpyxl")

    df.columns = [str(c).strip() for c in df.columns]

    ren = {}
    for i, c in enumerate(df.columns[:len(INDEX_TO_REQUIRED)]):
        target = INDEX_TO_REQUIRED.get(i)
        if target:
            ren[c] = target
    df = df.rename(columns=ren)

    for c in REQUIRED:
        if c not in df.columns:
            df[c] = None
    df = df[REQUIRED].copy()

    df["Tổng Sản lượng (Lẻ)"] = pd.to_numeric(df["Tổng Sản lượng (Lẻ)"], errors="coerce").fillna(0)
    df["Doanh số bán"] = pd.to_numeric(df["Doanh số bán"], errors="coerce").fillna(0)

    for c in ["Tên NPP", "Mã KH", "Tên KH", "Nhóm hàng", "Mã SP", "Tên SP"]:
        df[c] = df[c].astype(str).str.strip()

    return df

def build_pivot_by_customer(df: pd.DataFrame) -> pd.DataFrame:
    rep = (
        df.groupby("Mã KH")
          .agg(**{
              "Tên KH đại diện": ("Tên KH", _mode_text),
              "Tên NPP đại diện": ("Tên NPP", _mode_text),
          }).reset_index()
    )

    qty = pd.pivot_table(
        df, index=["Mã KH"], columns="Tên SP",
        values="Tổng Sản lượng (Lẻ)", aggfunc="sum", fill_value=0, observed=False
    )
    revenue = df.groupby("Mã KH", as_index=True)["Doanh số bán"].sum().to_frame("Tổng Doanh số")

    out = rep.set_index("Mã KH").join(qty, how="right").join(revenue, how="left").reset_index()

    fixed = ["Mã KH", "Tên KH đại diện", "Tên NPP đại diện"]
    cols = out.columns.tolist()
    dynamic = [c for c in cols if c not in fixed and c != "Tổng Doanh số"]
    if "Tổng Doanh số" in cols:
        dynamic.append("Tổng Doanh số")
    return out[fixed + dynamic]

# =========================
# BƯỚC 2 — PIVOT (STREAMING XLSX)
# =========================
def pivot_streaming_xlsx(xlsx_bytes: bytes, header_row_user: int = 1):
    """
    Đọc XLSX ở chế độ read_only, gom số theo KH/SKU, không tải toàn bộ vào RAM.
    Trả về (DataFrame pivot, số KH, số SKU).
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Đọc header theo dòng chỉ định
    headers_raw = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == header_row_user:
            headers_raw = list(row or [])
            break
    if headers_raw is None:
        raise ValueError("Không tìm thấy dòng tiêu đề theo 'Dòng tiêu đề'.")

    name2idx = {}
    norm_headers = [_normalize_header(h) for h in headers_raw]
    for req, aliases in NEEDED_NAMES.items():
        hit = None
        for j, nh in enumerate(norm_headers):
            if nh in aliases:
                hit = j; break
        if hit is not None:
            name2idx[req] = hit

    fallback_order = ["Tên NPP","Mã KH","Tên KH","Nhóm hàng","Mã SP","Tên SP","Tổng Sản lượng (Lẻ)","Doanh số bán"]
    for pos, col in enumerate(fallback_order):
        if col not in name2idx and pos < len(headers_raw):
            name2idx[col] = pos

    missing = [c for c in fallback_order if c not in name2idx]
    if missing:
        raise ValueError(f"Thiếu cột bắt buộc (không map được): {missing}")

    customers = {}
    all_products = set()

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= header_row_user:
            continue
        if row is None:
            continue

        def _get(colname):
            idx = name2idx[colname]
            return row[idx] if idx < len(row) else None

        ma_kh = str(_get("Mã KH") or "").strip()
        if not ma_kh:
            continue

        ten_kh = str(_get("Tên KH") or "").strip()
        ten_npp = str(_get("Tên NPP") or "").strip()
        ten_sp = str(_get("Tên SP") or "").strip()

        try:
            sl = float(_get("Tổng Sản lượng (Lẻ)") or 0)
        except Exception:
            sl = 0.0
        try:
            ds = float(_get("Doanh số bán") or 0)
        except Exception:
            ds = 0.0

        ent = customers.get(ma_kh)
        if ent is None:
            ent = {"kh_counts": Counter(),"npp_counts": Counter(),"revenue": 0.0,"products": Counter()}
            customers[ma_kh] = ent

        if ten_kh: ent["kh_counts"][ten_kh] += 1
        if ten_npp: ent["npp_counts"][ten_npp] += 1
        if ten_sp:
            ent["products"][ten_sp] += sl
            all_products.add(ten_sp)
        ent["revenue"] += ds

    prod_list = sorted(all_products)
    records = []
    for ma_kh, ent in customers.items():
        row = {
            "Mã KH": ma_kh,
            "Tên KH đại diện": _mode_text(list(ent["kh_counts"].elements())),
            "Tên NPP đại diện": _mode_text(list(ent["npp_counts"].elements())),
        }
        for p in prod_list:
            row[p] = int(ent["products"].get(p, 0))
        row["Tổng Doanh số"] = int(ent["revenue"])
        records.append(row)

    if not records:
        return pd.DataFrame(columns=["Mã KH","Tên KH đại diện","Tên NPP đại diện"] + prod_list + ["Tổng Doanh số"]), 0, 0

    df_pivot = pd.DataFrame.from_records(records)
    for p in prod_list:
        df_pivot[p] = pd.to_numeric(df_pivot[p], errors="coerce").fillna(0).astype(int)
    if "Tổng Doanh số" in df_pivot.columns:
        df_pivot["Tổng Doanh số"] = pd.to_numeric(df_pivot["Tổng Doanh số"], errors="coerce").fillna(0).astype(int)

    return df_pivot, len(customers), len(prod_list)

# =========================
# BƯỚC 2 — PIVOT (STREAMING CSV)
# =========================
def pivot_streaming_csv(csv_bytes: bytes, header_row_user: int = 1, delimiter=","):
    text = csv_bytes.decode("utf-8", errors="ignore").splitlines()
    reader = csv.reader(text, delimiter=delimiter)

    headers = None
    for i, row in enumerate(reader, start=1):
        if i == header_row_user:
            headers = row
            break
    if headers is None:
        raise ValueError("Không tìm thấy dòng tiêu đề trong CSV.")

    norm_headers = [_normalize_header(h) for h in headers]
    name2idx = {}
    for req, aliases in NEEDED_NAMES.items():
        hit = None
        for j, nh in enumerate(norm_headers):
            if nh in aliases:
                hit = j; break
        if hit is not None:
            name2idx[req] = hit
    fallback_order = ["Tên NPP","Mã KH","Tên KH","Nhóm hàng","Mã SP","Tên SP","Tổng Sản lượng (Lẻ)","Doanh số bán"]
    for pos, col in enumerate(fallback_order):
        if col not in name2idx and pos < len(headers):
            name2idx[col] = pos

    missing = [c for c in fallback_order if c not in name2idx]
    if missing:
        raise ValueError(f"Thiếu cột bắt buộc trong CSV: {missing}")

    customers = {}
    all_products = set()

    for row in reader:
        if not row:
            continue

        def _get(colname):
            idx = name2idx[colname]
            return row[idx] if idx < len(row) else ""

        ma_kh = (_get("Mã KH") or "").strip()
        if not ma_kh:
            continue
        ten_kh = (_get("Tên KH") or "").strip()
        ten_npp = (_get("Tên NPP") or "").strip()
        ten_sp = (_get("Tên SP") or "").strip()

        try:
            sl = float(_get("Tổng Sản lượng (Lẻ)") or 0)
        except:
            sl = 0.0
        try:
            ds = float(_get("Doanh số bán") or 0)
        except:
            ds = 0.0

        ent = customers.get(ma_kh)
        if ent is None:
            ent = {"kh_counts": Counter(),"npp_counts": Counter(),"revenue": 0.0,"products": Counter()}
            customers[ma_kh] = ent

        if ten_kh: ent["kh_counts"][ten_kh] += 1
        if ten_npp: ent["npp_counts"][ten_npp] += 1
        if ten_sp:
            ent["products"][ten_sp] += sl
            all_products.add(ten_sp)
        ent["revenue"] += ds

    prod_list = sorted(all_products)
    records = []
    for ma_kh, ent in customers.items():
        row = {
            "Mã KH": ma_kh,
            "Tên KH đại diện": _mode_text(list(ent["kh_counts"].elements())),
            "Tên NPP đại diện": _mode_text(list(ent["npp_counts"].elements())),
        }
        for p in prod_list:
            row[p] = int(ent["products"].get(p, 0))
        row["Tổng Doanh số"] = int(ent["revenue"])
        records.append(row)

    df_pivot = pd.DataFrame.from_records(records)
    for p in prod_list:
        df_pivot[p] = pd.to_numeric(df_pivot[p], errors="coerce").fillna(0).astype(int)
    if "Tổng Doanh số" in df_pivot.columns:
        df_pivot["Tổng Doanh số"] = pd.to_numeric(df_pivot["Tổng Doanh số"], errors="coerce").fillna(0).astype(int)
    return df_pivot, len(customers), len(prod_list)

# =========================
# UI — BƯỚC 1: CẮT FILE
# =========================
st.header("Bước 1 — ✂️ Cắt cột từ file nặng (pandas, giữ nguyên số dòng)")
with st.expander("Chọn & cắt file", expanded=True):
    c1, c2 = st.columns([2, 1])
    with c1:
        raw_file = st.file_uploader(
            "Upload file Excel GỐC (xlsx/xlsm/xls; nếu xlsb hãy lưu lại .xlsx trước khi dùng)",
            type=["xlsx", "xlsm", "xls"],
            key="raw_file",
        )
    with c2:
        sheet_name = None
        if raw_file:
            try:
                xls = pd.ExcelFile(raw_file, engine="openpyxl")
                sheet_name = st.selectbox(
                    "Chọn sheet cần cắt",
                    xls.sheet_names,
                    index=0,
                    key="cut_sheet",
                )
            except Exception as e:
                st.warning(f"Không đọc được danh sách sheet (sẽ dùng sheet đầu tiên). Chi tiết: {e}")
                sheet_name = None

    if raw_file and st.button("✂️ CẮT NGAY", key="btn_cut", use_container_width=True):
        try:
            with st.spinner("Đang cắt cột bằng pandas..."):
                trimmed_bytes = cut_with_pandas_keep_rows(
                    file_bytes=raw_file.getvalue(),
                    sheet_name=sheet_name,
                )
            st.session_state["trimmed_bytes"] = trimmed_bytes
            st.success("✅ Đã cắt xong. Tải về hoặc dùng trực tiếp cho Bước 2.")
            st.download_button(
                "⬇️ Tải file đã cắt (.xlsx)",
                data=trimmed_bytes,
                file_name=f"{os.path.splitext(raw_file.name)[0]}_filtered_preserve.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Lỗi khi cắt file: {e}")

# =========================
# UI — BƯỚC 2: PIVOT
# =========================
st.header("Bước 2 — 📊 Pivot sản lượng & doanh số theo Khách Hàng")

src_choice = st.radio(
    "Nguồn file ‘đã cắt’ để pivot:",
    ["Dùng file đã cắt ở Bước 1", "Upload file đã cắt (XLSX)"],
    horizontal=True,
    index=0,
)

trimmed_to_use = None
if src_choice == "Dùng file đã cắt ở Bước 1":
    if "trimmed_bytes" in st.session_state:
        trimmed_to_use = st.session_state["trimmed_bytes"]
    else:
        st.info("Chưa có file đã cắt trong session. Hãy thực hiện Bước 1 hoặc chọn 'Upload file đã cắt (XLSX)'.")
else:
    up2 = st.file_uploader("Upload file ĐÃ CẮT (XLSX 8 cột)", type=["xlsx"], key="trimmed_upload")
    if up2:
        trimmed_to_use = up2.getvalue()

pivot_mode = st.radio(
    "Chế độ Pivot:",
    ["Nhanh (pandas đọc XLSX)", "Siêu nhẹ (streaming XLSX)", "Siêu nhẹ (CSV)"],
    horizontal=True,
    key="pivot_mode",
)

header_row_user = st.number_input(
    "Dòng tiêu đề (1 = dòng đầu)",
    min_value=1, value=1, step=1, key="pivot_header_row"
)

if pivot_mode == "Nhanh (pandas đọc XLSX)":
    if trimmed_to_use and st.button("🚀 PIVOT (pandas)", use_container_width=True, key="btn_pivot_pd"):
        try:
            with st.spinner("Đang đọc & chuẩn hoá (pandas)..."):
                df_cut = pd.read_excel(BytesIO(trimmed_to_use), header=0, engine="openpyxl")
                df_norm = normalize_after_cut(df_cut, header_row_user=header_row_user)

                miss = [c for c in REQUIRED if c not in df_norm.columns]
                if miss:
                    st.error(f"Thiếu cột bắt buộc: {miss}")
                    st.stop()

                pivot_df = build_pivot_by_customer(df_norm)

            st.success("✅ Hoàn tất (pandas)!")
            st.dataframe(pivot_df, use_container_width=True)

            out = BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as w:
                pivot_df.to_excel(w, index=False, sheet_name="PIVOT_KH")
            st.download_button(
                "⬇️ Tải Excel PIVOT",
                data=out.getvalue(),
                file_name="pivot_sku_theo_khachhang.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Lỗi khi pivot (pandas): {e}")

elif pivot_mode == "Siêu nhẹ (streaming XLSX)":
    if trimmed_to_use and st.button("🚀 PIVOT (streaming XLSX)", use_container_width=True, key="btn_pivot_stream_xlsx"):
        try:
            with st.spinner("Đang pivot streaming XLSX..."):
                pivot_df, n_cust, n_sku = pivot_streaming_xlsx(trimmed_to_use, header_row_user=header_row_user)
            st.success(f"✅ Xong! KH: {n_cust:,} • SKU: {n_sku:,}")
            st.dataframe(pivot_df, use_container_width=True)

            out = BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as w:
                pivot_df.to_excel(w, index=False, sheet_name="PIVOT_KH")
            st.download_button(
                "⬇️ Tải Excel PIVOT",
                data=out.getvalue(),
                file_name="pivot_streaming_xlsx.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Lỗi pivot (streaming XLSX): {e}")

elif pivot_mode == "Siêu nhẹ (CSV)":
    csv_file = st.file_uploader("Upload CSV (khuyến nghị: Save As từ Excel → CSV UTF‑8)", type=["csv"], key="csv_upload")
    if csv_file and st.button("🚀 PIVOT (CSV streaming)", use_container_width=True, key="btn_pivot_stream_csv"):
        try:
            with st.spinner("Đang pivot streaming CSV..."):
                pivot_df, n_cust, n_sku = pivot_streaming_csv(csv_file.getvalue(), header_row_user=header_row_user)
            st.success(f"✅ Xong! KH: {n_cust:,} • SKU: {n_sku:,}")
            st.dataframe(pivot_df, use_container_width=True)

            out = BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as w:
                pivot_df.to_excel(w, index=False, sheet_name="PIVOT_KH")
            st.download_button(
                "⬇️ Tải Excel PIVOT",
                data=out.getvalue(),
                file_name="pivot_streaming_csv.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Lỗi pivot (CSV streaming): {e}")
